import requests
from openpyxl import load_workbook, Workbook
import datetime
import tkinter as tk
from tkinter import messagebox
import webbrowser

# Configurações
API_KEY = "d647b955ad05576e80db1a9b4ab86b35"  # Substitua pela sua chave da OpenWeatherMap
CIDADE = "São Paulo, BR"  # Altere para sua cidade
ARQUIVO_EXCEL = "historico_temperatura.xlsx"
URL_PREVISAO = "https://www.google.com/search?q=previs%C3%A3o+do+tempo+s%C3%A3o+paulo&sca_esv=bf6563b3cfaca2c5&sxsrf=AHTn8zpqSQIrygJITJiW7gi980bsEDfriA%3A1743190033859&ei=EfjmZ6eYNMv-5OUP_o7gkAg&oq=previs%C3%A3o+do+tempo+s%C3%A3o+paulo&gs_lp=Egxnd3Mtd2l6LXNlcnAiHXByZXZpc8OjbyBkbyB0ZW1wbyBzw6NvIHBhdWxvKgIIADIKECMYgAQYJxiKBTIFEAAYgAQyBRAAGIAEMgUQABiABDIFEAAYgAQyBRAAGIAEMgUQABiABDIFEAAYgAQyBRAAGIAEMgUQABiABEjBIVCdA1iSGHABeAGQAQCYAb0BoAGnCaoBBDAuMTC4AQHIAQD4AQGYAgugAugKwgIHECMYsAMYJ8ICChAAGLADGNYEGEfCAhMQLhiABBiwAxhDGMgDGIoF2AEBwgIEECMYJ8ICCBAAGIAEGLEDmAMAiAYBkAYLugYECAEYCJIHBTEuOS4xoAfqTrIHBTAuOS4xuAe-Cg&sclient=gws-wiz-serp"

# Função para obter a temperatura e umidade via API
def capturar_temperatura():
    try:
        webbrowser.open(URL_PREVISAO)  # Abre a página da previsão do tempo

        url = f"http://api.openweathermap.org/data/2.5/weather?q={CIDADE}&appid={API_KEY}&units=metric&lang=pt"
        resposta = requests.get(url)
        dados = resposta.json()

        if resposta.status_code == 200:
            temperatura = dados["main"]["temp"]
            umidade = dados["main"]["humidity"]  # Obtém a umidade
            agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            salvar_dados(agora, temperatura, umidade)  # Inclui a umidade na chamada
            messagebox.showinfo("Sucesso", f"Temperatura: {temperatura}°C\nUmidade: {umidade}%\nHorário: {agora}\nSalvo no Excel!")
        else:
            messagebox.showerror("Erro", f"Erro na API: {dados['message']}")

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao capturar temperatura: {e}")

# Função para salvar os dados no Excel
def salvar_dados(horario, temperatura, umidade):  # Adiciona o parâmetro umidade
    try:
        wb = load_workbook(ARQUIVO_EXCEL)
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Data e Hora", "Temperatura (°C)", "Umidade (%)"])  # Adiciona a coluna de umidade

    # Verifica se a planilha está vazia (apenas com o cabeçalho)
    if sheet.max_row == 1:
        linha = 2  # Começa na segunda linha (abaixo do cabeçalho)
    else:
        linha = sheet.max_row + 1  # Adiciona na próxima linha disponível

    sheet.cell(row=linha, column=1, value=horario)
    sheet.cell(row=linha, column=2, value=temperatura)
    sheet.cell(row=linha, column=3, value=f"{umidade}%")  # Inclui a umidade na linha com o símbolo %
    wb.save(ARQUIVO_EXCEL)

# Criando interface gráfica com Tkinter
root = tk.Tk()
root.title("Captura de Temperatura e Umidade")

label = tk.Label(root, text="Clique no botão para capturar a temperatura e umidade atual:")
label.pack(pady=10)

botao = tk.Button(root, text="Capturar Dados", command=capturar_temperatura)
botao.pack(pady=10)

root.mainloop()